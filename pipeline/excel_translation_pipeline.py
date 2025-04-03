import os
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import range_boundaries
import json
from datetime import datetime
from .skip_pipeline import should_translate
from config.log_config import app_logger


def extract_excel_content_to_json(file_path):
    workbook = load_workbook(file_path)
    cell_data = []
    count = 0
    
    # Add sheet names to the extraction process
    for sheet_name in workbook.sheetnames:
        # Add sheet name as a special entry if it should be translated
        if should_translate(sheet_name):
            count += 1
            sheet_info = {
                "count": count,
                "sheet": "SHEET_NAME",  # Special marker to identify sheet names
                "row": 0,               # Use 0 to indicate it's a sheet name, not a cell
                "column": 0,            # Use 0 to indicate it's a sheet name, not a cell
                "value": sheet_name,
                "is_merged": False,
                "is_sheet_name": True   # Flag to identify this as a sheet name entry
            }
            cell_data.append(sheet_info)
        
        sheet = workbook[sheet_name]
        merged_cells_ranges = sheet.merged_cells.ranges

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None or isinstance(cell.value, datetime) or not should_translate(str(cell.value)):
                    continue
                if isinstance(cell, MergedCell):
                    continue
                is_merged_cell = False
                for merged_range in merged_cells_ranges:
                    min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                    if cell.row == min_row and cell.column == min_col:
                        is_merged_cell = True
                        break
                # Convert datetime values to string
                cell_value = str(cell.value).replace("\n", "␊").replace("\r", "␍")
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.isoformat()
                count += 1
                cell_info = {
                    "count": count,
                    "sheet": sheet_name,
                    "row": cell.row,
                    "column": cell.column,
                    "value": cell_value,
                    "is_merged": is_merged_cell,
                    "is_sheet_name": False  # Regular cell, not a sheet name
                }
                cell_data.append(cell_info)
    
    filename = os.path.splitext(os.path.basename(file_path))[0]
    temp_folder = os.path.join("temp", filename)
    os.makedirs(temp_folder, exist_ok=True)
    json_path = os.path.join(temp_folder, "src.json")
    
    with open(json_path, "w", encoding="utf-8") as json_file:
        json.dump(cell_data, json_file, ensure_ascii=False, indent=4)

    return json_path

def write_translated_content_to_excel(file_path, original_json_path, translated_json_path):
    workbook = load_workbook(file_path)

    # Load original JSON data
    with open(original_json_path, "r", encoding="utf-8") as original_file:
        original_data = json.load(original_file)
    
    # Load translated JSON data
    with open(translated_json_path, "r", encoding="utf-8") as translated_file:
        translated_data = json.load(translated_file)

    # Convert translations to a dictionary {count: translated_value}
    translations = {str(item["count"]): item["translated"] for item in translated_data}
    
    # Track sheet name translations to apply at the end
    sheet_name_translations = {}
    
    # First pass: Collect sheet name translations
    for cell_info in original_data:
        count = str(cell_info["count"])  # Ensure count is a string
        if cell_info.get("is_sheet_name", False):
            original_sheet_name = cell_info["value"]
            translated_sheet_name = translations.get(count)
            if translated_sheet_name:
                sheet_name_translations[original_sheet_name] = translated_sheet_name.replace("␊", "\n").replace("␍", "\r")
    
    # Second pass: Update cell contents
    for cell_info in original_data:
        # Skip sheet name entries as they are handled separately
        if cell_info.get("is_sheet_name", False):
            continue
            
        count = str(cell_info["count"])  # Ensure count is a string
        sheet_name = cell_info["sheet"]
        row = cell_info["row"]
        column = cell_info["column"]
        original_text = cell_info["value"]
        is_merged = cell_info.get("is_merged", False)

        # Get the translated text
        value = translations.get(count, None)
        if value is None:
            # Log missing translation with original text
            app_logger.warning(
                f"Translation missing for count {count}. Original text: '{original_text}'"
            )
            continue
        
        # Replace line breaks to preserve format
        value = value.replace("␊", "\n").replace("␍", "\r")

        # Write to the Excel cell
        sheet = workbook[sheet_name]
        cell = sheet.cell(row=row, column=column)
        cell.value = value

        # Handle merged cells if applicable
        if is_merged:
            merge_range = f"{cell.coordinate}:{cell.coordinate}"
            sheet.merge_cells(merge_range)
    
    # Final pass: Rename sheets with their translations
    for original_name, translated_name in sheet_name_translations.items():
        if original_name in workbook.sheetnames:
            sheet = workbook[original_name]
            sheet.title = translated_name
            app_logger.info(f"Renamed sheet from '{original_name}' to '{translated_name}'")

    # Save the modified Excel file
    result_folder = os.path.join('result')
    os.makedirs(result_folder, exist_ok=True)
    
    result_path = os.path.join(
        result_folder,
        f"{os.path.splitext(os.path.basename(file_path))[0]}_translated{os.path.splitext(file_path)[1]}"
    )
    
    workbook.save(result_path)
    app_logger.info(f"Translated Excel saved to: {result_path}")
    return result_path